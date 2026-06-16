import os
import json
import re
import urllib.parse
import io
import requests
from flask import Flask, request, jsonify, render_template, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__, template_folder='templates', static_folder='static')

# Configuration
DATA_FILE = os.path.join(os.path.dirname(__file__), 'leads_store.json')
MOCK_FILE = os.path.join(os.path.dirname(__file__), 'mock_leads.json')

def load_leads():
    """Loads leads from local store, fallback to mock if store doesn't exist."""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    
    if os.path.exists(MOCK_FILE):
        try:
            with open(MOCK_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                save_leads(data)
                return data
        except Exception:
            pass
            
    return []

def save_leads(leads):
    """Saves leads to local store."""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(leads, f, ensure_ascii=False, indent=2)

def rule_based_scorer(requirement):
    """Fallback rule-based lead scoring logic matching tieu_chi_cham_diem.txt."""
    req_lower = requirement.lower()
    score = 0
    reason_parts = []
    
    # 1. Parse budget in billions (tỷ)
    budget_val = None
    # Matches patterns like: 35 tỷ, 35.5 tỷ, 1,5 tỷ, 1-2 tỷ, 3-4 tỷ
    matches = re.findall(r'(\d+(?:[.,]\d+)?)\s*(?:-|đến)?\s*(\d+(?:[.,]\d+)?)\s*(?:tỷ|ty)', req_lower)
    if matches:
        for match in matches:
            val1 = float(match[0].replace(',', '.'))
            val2 = float(match[1].replace(',', '.'))
            budget_val = max(val1, val2)
    else:
        single_matches = re.findall(r'(\d+(?:[.,]\d+)?)\s*(?:tỷ|ty)', req_lower)
        if single_matches:
            budget_val = max([float(m.replace(',', '.')) for m in single_matches])
            
    # 2. Check VIP Criteria (+50)
    vip_detected = False
    vip_reasons = []
    
    if budget_val is not None and budget_val >= 20.0:
        vip_detected = True
        vip_reasons.append(f"Ngân sách lớn ({budget_val} tỷ)")
    elif "tài chính mạnh" in req_lower or "tai chinh manh" in req_lower or "không thành vấn đề" in req_lower or "khong thanh van de" in req_lower:
        vip_detected = True
        vip_reasons.append("Tài chính mạnh")
        
    premium_types = ["biệt thự", "biet thu", "penthouse", "shophouse mặt đường", "shophouse mat duong", "đất công nghiệp", "dat cong nghiep", "sàn văn phòng diện tích lớn", "san van phong dien tich lon", "sàn văn phòng", "san van phong"]
    for pt in premium_types:
        if pt in req_lower:
            vip_detected = True
            vip_reasons.append(f"Loại hình cao cấp ({pt})")
            break
            
    prime_locations = ["quận 1", "quan 1", "ven sông", "ven song", "ocean park", "phú mỹ hưng", "phu my hung"]
    has_low_budget = (budget_val is not None and budget_val <= 2.0)
    if not has_low_budget:
        for pl in prime_locations:
            if pl in req_lower:
                vip_detected = True
                vip_reasons.append(f"Vị trí đắc địa ({pl})")
                break
            
    profiles = ["chủ doanh nghiệp", "chu doanh nghiep", "nhà đầu tư chuyên nghiệp", "nha dau tu chuyen nghiep", "mua sỉ", "mua si", "số lượng lớn", "so luong lon"]
    for pr in profiles:
        if pr in req_lower:
            vip_detected = True
            vip_reasons.append(f"Đối tượng VIP ({pr})")
            break
            
    urgency = ["pháp lý chuẩn", "phap ly chuan", "sổ hồng riêng", "so hong rieng", "gặp trực tiếp chủ đầu tư", "gap truc tiep chu dau tu"]
    for urg in urgency:
        if urg in req_lower:
            vip_detected = True
            vip_reasons.append(f"Tính cấp thiết & Minh bạch ({urg})")
            break
            
    # 3. Check Trash Criteria (-50)
    trash_detected = False
    trash_reasons = []
    
    # Unrealistic budget check
    has_central = any(x in req_lower for x in ["quận 1", "quan 1", "trung tâm", "trung tam"])
    if has_central and budget_val is not None and budget_val <= 2.0:
        trash_detected = True
        trash_reasons.append(f"Yêu cầu phi thực tế (nhà trung tâm giá {budget_val} tỷ)")
    elif "vài trăm triệu" in req_lower or "vai tram trieu" in req_lower:
        trash_detected = True
        trash_reasons.append("Yêu cầu phi thực tế (biệt thự/nhà trung tâm vài trăm triệu)")
        
    no_demands = ["nhầm số", "nham so", "không có nhu cầu", "khong co nhu cau", "dữ liệu cũ", "du lieu cu", "nhầm ngành", "nham nganh", "nhầm người", "nham nguoi"]
    for nd in no_demands:
        if nd in req_lower:
            trash_detected = True
            trash_reasons.append(f"Không có nhu cầu ({nd})")
            break
            
    uncooperative = ["hỏi giá cho vui", "hoi gia cho vui", "chưa có ý định mua", "chua co y dinh mua", "không thiện chí", "thái độ không hợp tác", "khong hop tac"]
    for uc in uncooperative:
        if uc in req_lower:
            trash_detected = True
            trash_reasons.append(f"Khách hàng không thiện chí ({uc})")
            break
            
    # Check Spam/Advertising (only if no buyer intent found)
    has_buyer_intent = any(x in req_lower for x in ["cần mua", "can mua", "tìm mua", "tim mua", "mua chung cư", "mua chung cu", "mua nhà", "mua nha", "tìm căn", "tim can", "nhu cầu tìm"])
    if not has_buyer_intent:
        spam_services = ["bảo hiểm", "bao hiem", "dịch vụ vay", "dich vu vay", "cho vay", "mời chào dịch vụ", "moi chao dich vu", "chuyên cung cấp", "chuyen cung cap"]
        for ss in spam_services:
            if ss in req_lower:
                trash_detected = True
                trash_reasons.append(f"Spam/Quảng cáo dịch vụ ({ss})")
                break
                
    comm_errors = ["thuê bao", "thue bao", "không bắt máy", "khong bat may", "không phản hồi zalo", "khong phan hoi zalo", "không liên lạc được", "khong lien lac duoc"]
    for ce in comm_errors:
        if ce in req_lower:
            trash_detected = True
            trash_reasons.append(f"Lỗi liên lạc ({ce})")
            break
            
    # 4. Final scoring assembly
    if trash_detected:
        score = -50
        classification = "Rác"
        reason = "Trừ 50đ: " + ", ".join(trash_reasons)
    elif vip_detected:
        score = 50
        classification = "VIP"
        reason = "Cộng 50đ: " + ", ".join(vip_reasons)
    else:
        score = 0
        classification = "Trung bình"
        reason = "Khách hàng tầm trung, có nhu cầu thực tế liên quan đến chung cư/nhà phố hoặc cần tư vấn thêm."
        
    return {
        "score": score,
        "classification": classification,
        "reason": f"[Từ khóa] {reason}"
    }

def call_gemini_api(api_key, requirement):
    """Calls Gemini API using standard HTTP requests."""
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}"
    
    prompt = f"""
Bạn là AI chấm điểm khách hàng tiềm năng bất động sản dựa trên quy tắc sau:

1. TIÊU CHÍ CỘNG 50 ĐIỂM (KHÁCH HÀNG VIP/SIÊU TIỀM NĂNG)
- Ngân sách lớn: Có đề cập đến số tiền từ 20 tỷ trở lên hoặc "tài chính mạnh", "không thành vấn đề".
- Loại hình cao cấp: "Biệt thự đơn lập", "Penthouse", "Shophouse mặt đường lớn", "Quỹ đất công nghiệp", "Sàn văn phòng diện tích lớn".
- Vị trí đắc địa: "Quận 1", "Ven sông", "Vinhomes Ocean Park", "Phú Mỹ Hưng".
- Đối tượng khách hàng: "Chủ doanh nghiệp", "Nhà đầu tư chuyên nghiệp", "Mua sỉ", "Mua số lượng lớn".
- Tính cấp thiết & Minh bạch: "Pháp lý chuẩn 100%", "Sổ hồng riêng", "Muốn gặp trực tiếp chủ đầu tư để đàm phán".

2. TIÊU CHÍ TRỪ 50 ĐIỂM (KHÁCH HÀNG RÁC/KHÔNG TIỀM NĂNG)
- Yêu cầu phi thực tế: Mua bds giá thấp vô lý (VD: Nhà Q1 giá 1-2 tỷ, biệt thự có hồ bơi giá vài trăm triệu).
- Không có nhu cầu: "Nhầm số", "Không có nhu cầu", "Dữ liệu cũ", "Nhầm ngành".
- Khách hàng không thiện chí: "Hỏi giá cho vui", "Chưa có ý định mua", "Thái độ không hợp tác".
- Spam/Quảng cáo: "Bảo hiểm", "Vay vốn", "Mời chào dịch vụ".
- Thông tin liên lạc lỗi: "Thuê bao", "Gọi nhiều lần không bắt máy", "Không phản hồi Zalo".

3. CÁC TRƯỜNG HỢP KHÁC (GIỮ NGUYÊN ĐIỂM 0 HOẶC CỘNG NHẸ 10 ĐIỂM)
- Nhu cầu thực chung cư, nhà phố tầm trung (3-10 tỷ).
- Cần vay ngân hàng, đang cân nhắc chính sách.
- Cần tư vấn thêm pháp lý hoặc vị trí.

Hãy phân tích nhu cầu sau và trả về kết quả dưới dạng JSON (không markdown, không bao quanh bởi ```json):
Nhu cầu: "{requirement}"

JSON Format:
{{
  "score": số_điểm_đánh_giá,
  "classification": "VIP" hoặc "Trung bình" hoặc "Rác",
  "reason": "Giải thích lý do bằng tiếng Việt ngắn gọn dựa trên quy tắc"
}}
"""

    headers = {
        "Content-Type": "application/json"
    }
    
    payload = {
        "contents": [{
            "parts": [{
                "text": prompt
            }]
        }],
        "generationConfig": {
            "responseMimeType": "application/json"
        }
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=15)
        if response.status_code == 200:
            res_json = response.json()
            text_out = res_json['candidates'][0]['content']['parts'][0]['text']
            
            # Clean text if model ignored responseMimeType
            text_out = text_out.strip()
            if text_out.startswith("```json"):
                text_out = text_out[7:]
            if text_out.endswith("```"):
                text_out = text_out[:-3]
            text_out = text_out.strip()
            
            parsed = json.loads(text_out)
            return {
                "score": int(parsed.get("score", 0)),
                "classification": str(parsed.get("classification", "Trung bình")),
                "reason": f"[AI Gemini] {str(parsed.get('reason', 'Đã chấm điểm'))}"
            }
        else:
            return {
                "error": True,
                "msg": f"API Gemini trả về mã lỗi: {response.status_code}"
            }
    except Exception as e:
        return {
            "error": True,
            "msg": f"Lỗi gọi API Gemini: {str(e)}"
        }

def clean_google_sheet_url(url):
    """Converts a standard Google Sheet edit link to export CSV link."""
    # Pattern to match: https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit...
    match = re.search(r'\/spreadsheets\/d\/([a-zA-Z0-9-_]+)', url)
    if not match:
        return url
    
    sheet_id = match.group(1)
    gid = "0"
    
    # Extract gid from URL parameters
    gid_match = re.search(r'gid=(\d+)', url)
    if gid_match:
        gid = gid_match.group(1)
        
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

def normalize_dataframe(df):
    """Maps custom columns of DataFrame to standard lead columns."""
    cols = [str(c).lower().strip() for c in df.columns]
    
    # Header Mapping
    name_cols = ['họ tên', 'ho ten', 'họ và tên', 'ho va ten', 'tên', 'ten', 'name', 'khách hàng', 'khach hang', 'kh']
    phone_cols = ['số điện thoại', 'so dien thoai', 'sđt', 'sdt', 'phone', 'điện thoại', 'dien thoai', 'tel', 'mobile']
    email_cols = ['email', 'mail', 'thư điện tử', 'thu dien tu']
    req_cols = ['nhu cầu chi tiết', 'nhu cau chi tiet', 'nhu cầu', 'nhu cau', 'chi tiết', 'chi tiet', 'nội dung', 'noi dung', 'ghi chú', 'ghi chu', 'notes', 'mô tả', 'mo ta']
    
    mapped_name = None
    mapped_phone = None
    mapped_email = None
    mapped_req = None
    
    for idx, c in enumerate(cols):
        orig_col = df.columns[idx]
        if c in name_cols and not mapped_name:
            mapped_name = orig_col
        elif c in phone_cols and not mapped_phone:
            mapped_phone = orig_col
        elif c in email_cols and not mapped_email:
            mapped_email = orig_col
        elif c in req_cols and not mapped_req:
            mapped_req = orig_col
            
    # Try positional fallback if mapping failed
    if not mapped_req:
        # Usually requirement is a long text column, we can guess
        for col in df.columns:
            if df[col].dtype == object and df[col].str.len().mean() > 30:
                mapped_req = col
                break
        if not mapped_req and len(df.columns) >= 3:
            mapped_req = df.columns[2] # Fallback to 3rd col
            
    if not mapped_name and len(df.columns) >= 1:
        mapped_name = df.columns[0]
        
    if not mapped_phone and len(df.columns) >= 2:
        mapped_phone = df.columns[1]

    # Construct clean lead records
    leads = []
    for idx, row in df.iterrows():
        name = str(row[mapped_name]).strip() if mapped_name and pd.notna(row[mapped_name]) else "Khách hàng mới"
        phone = str(row[mapped_phone]).strip() if mapped_phone and pd.notna(row[mapped_phone]) else ""
        email = str(row[mapped_email]).strip() if mapped_email and pd.notna(row[mapped_email]) else ""
        requirement = str(row[mapped_req]).strip() if mapped_req and pd.notna(row[mapped_req]) else ""
        
        # Skip header rows that get accidentally parsed, or empty rows
        if not phone and not requirement:
            continue
            
        leads.append({
            "id": idx + 1,
            "name": name,
            "phone": phone,
            "email": email,
            "requirement": requirement,
            "score": None,
            "classification": None,
            "reason": None,
            "status": "Chưa duyệt",
            "reviewer_notes": ""
        })
    return leads

# --- API ROUTES ---

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/leads', methods=['GET'])
def get_leads_api():
    return jsonify(load_leads())

@app.route('/api/load-mock', methods=['POST'])
def load_mock_api():
    if os.path.exists(MOCK_FILE):
        try:
            with open(MOCK_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                save_leads(data)
                return jsonify({"status": "success", "msg": "Đã tải bộ dữ liệu mẫu thành công!", "data": data})
        except Exception as e:
            return jsonify({"status": "error", "msg": f"Không đọc được file mock: {str(e)}"}), 500
    return jsonify({"status": "error", "msg": "Không tìm thấy file mock_leads.json"}), 404

@app.route('/api/leads/update', methods=['POST'])
def update_lead_api():
    data = request.json
    if not data or 'id' not in data:
        return jsonify({"status": "error", "msg": "Thiếu thông tin ID khách hàng"}), 400
        
    leads = load_leads()
    lead_id = int(data['id'])
    
    found = False
    for lead in leads:
        if lead['id'] == lead_id:
            lead['name'] = data.get('name', lead.get('name'))
            lead['phone'] = data.get('phone', lead.get('phone'))
            lead['email'] = data.get('email', lead.get('email'))
            lead['requirement'] = data.get('requirement', lead.get('requirement'))
            
            # Allow manual override of scoring
            if 'score' in data and data['score'] is not None:
                lead['score'] = int(data['score'])
            if 'classification' in data:
                lead['classification'] = data['classification']
            if 'reason' in data:
                lead['reason'] = data['reason']
                
            lead['status'] = data.get('status', lead.get('status', 'Chưa duyệt'))
            lead['reviewer_notes'] = data.get('reviewer_notes', lead.get('reviewer_notes', ''))
            found = True
            break
            
    if not found:
        return jsonify({"status": "error", "msg": "Không tìm thấy khách hàng này"}), 404
        
    save_leads(leads)
    return jsonify({"status": "success", "msg": "Đã cập nhật thông tin khách hàng!", "data": leads})

@app.route('/api/fetch-sheet', methods=['POST'])
def fetch_sheet_api():
    req_data = request.json
    url = req_data.get('url')
    if not url:
        return jsonify({"status": "error", "msg": "Vui lòng cung cấp link Google Sheet"}), 400
        
    csv_url = clean_google_sheet_url(url)
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        res = requests.get(csv_url, headers=headers, timeout=15)
        if res.status_code == 404:
            return jsonify({
                "status": "error", 
                "msg": "Không thể truy cập Google Sheet. Tệp có thể ở chế độ riêng tư (404). Vui lòng kiểm tra quyền chia sẻ 'Người có liên kết có thể xem' hoặc tải lên tệp CSV/Excel trực tiếp bên dưới."
            }), 400
            
        if res.status_code != 200:
            return jsonify({"status": "error", "msg": f"Lỗi tải trang: Mã phản hồi {res.status_code}"}), 400
            
        # Try parsing as CSV
        df = pd.read_csv(io.StringIO(res.text))
        leads = normalize_dataframe(df)
        save_leads(leads)
        return jsonify({"status": "success", "msg": f"Đã kéo thành công {len(leads)} khách hàng từ Google Sheets!", "data": leads})
        
    except Exception as e:
        return jsonify({"status": "error", "msg": f"Lỗi xử lý file Google Sheet: {str(e)}"}), 500

@app.route('/api/upload-file', methods=['POST'])
def upload_file_api():
    if 'file' not in request.files:
        return jsonify({"status": "error", "msg": "Không tìm thấy tệp tải lên"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "msg": "Tên tệp rỗng"}), 400
        
    try:
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif filename.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file)
        else:
            return jsonify({"status": "error", "msg": "Định dạng tệp không được hỗ trợ. Hãy dùng .csv hoặc .xlsx"}), 400
            
        leads = normalize_dataframe(df)
        save_leads(leads)
        return jsonify({"status": "success", "msg": f"Đã nạp thành công {len(leads)} khách hàng từ tệp!", "data": leads})
    except Exception as e:
        return jsonify({"status": "error", "msg": f"Lỗi phân tích tệp tải lên: {str(e)}"}), 500

@app.route('/api/score', methods=['POST'])
def score_leads_api():
    req_data = request.json or {}
    api_key = req_data.get('api_key', '').strip()
    lead_ids = req_data.get('ids', []) # optional: list of specific IDs to re-score
    
    leads = load_leads()
    if not leads:
        return jsonify({"status": "error", "msg": "Không có khách hàng nào để chấm điểm"}), 400
        
    updated_count = 0
    errors = []
    
    for lead in leads:
        # If specific IDs are requested, skip others
        if lead_ids and lead['id'] not in lead_ids:
            continue
            
        req = lead['requirement']
        if not req:
            lead['score'] = 0
            lead['classification'] = "Trung bình"
            lead['reason'] = "Nhu cầu để trống"
            updated_count += 1
            continue
            
        # Run evaluation
        if api_key:
            res = call_gemini_api(api_key, req)
            if 'error' in res:
                # Add warning and fallback to rule-based
                errors.append(res['msg'])
                res = rule_based_scorer(req)
                res['reason'] = f"{res['reason']} (AI Lỗi: fallback từ khóa)"
        else:
            res = rule_based_scorer(req)
            
        lead['score'] = res['score']
        lead['classification'] = res['classification']
        lead['reason'] = res['reason']
        updated_count += 1
        
    save_leads(leads)
    
    msg = f"Đã chấm điểm xong {updated_count} khách hàng!"
    if errors:
        msg += f" (Phát hiện {len(errors)} lỗi kết nối AI, hệ thống đã tự động chạy bộ lọc từ khóa fallback)"
        
    return jsonify({
        "status": "success", 
        "msg": msg, 
        "data": leads,
        "warnings": errors
    })

@app.route('/api/export', methods=['GET'])
def export_leads_api():
    leads = load_leads()
    if not leads:
        # Create an empty template
        leads = [{
            "id": 1,
            "name": "Mẫu Nguyễn Văn A",
            "phone": "0900111222",
            "email": "a@example.com",
            "requirement": "Cần tìm căn hộ chung cư 3 tỷ",
            "score": 0,
            "classification": "Trung bình",
            "reason": "Phân loại mặc định",
            "status": "Chưa duyệt",
            "reviewer_notes": ""
        }]
        
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo Leads"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Mã Lead", "Họ Tên Khách Hàng", "Số Điện Thoại", "Email", 
        "Nhu Cầu Chi Tiết", "Điểm Đánh Giá", "Phân Loại AI", 
        "Lý Do Chấm Điểm", "Trạng Thái Duyệt", "Ghi Chú Kiểm Duyệt"
    ]
    ws.append(headers)
    
    # Design Styles
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Blue
    
    data_font = Font(name=font_family, size=11)
    
    # Border
    thin = Side(border_style="thin", color="D9D9D9")
    double = Side(border_style="double", color="1F4E79")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Center & Left alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Fills for Classification
    fill_vip = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Gold
    fill_med = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Soft Blue
    fill_trash = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red
    
    # Fills for Human Status
    fill_approved = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
    fill_rejected = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red
    fill_pending = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Yellow
    
    # Format Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = Border(left=thin, right=thin, top=thin, bottom=double)
    ws.row_dimensions[1].height = 28
    
    # Populate Data
    for lead in leads:
        row_data = [
            f"L{lead.get('id', 0):03d}",
            lead.get('name', ''),
            lead.get('phone', ''),
            lead.get('email', ''),
            lead.get('requirement', ''),
            lead.get('score', 0) if lead.get('score') is not None else "",
            lead.get('classification', 'Chưa phân loại'),
            lead.get('reason', ''),
            lead.get('status', 'Chưa duyệt'),
            lead.get('reviewer_notes', '')
        ]
        ws.append(row_data)
        
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = 24
        
        # Format Data Row cells
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = grid_border
            cell.alignment = align_left
            
            # Align center for codes, numbers, and statuses
            if col_idx in [1, 3, 6, 7, 9]:
                cell.alignment = align_center
                
            # Classifications Fills
            if col_idx == 7: # Phân loại AI
                val = str(cell.value)
                if val == "VIP":
                    cell.fill = fill_vip
                    cell.font = Font(name=font_family, size=11, bold=True, color="7F6000")
                elif val == "Trung bình":
                    cell.fill = fill_med
                    cell.font = Font(name=font_family, size=11, bold=False, color="1F4E79")
                elif val == "Rác":
                    cell.fill = fill_trash
                    cell.font = Font(name=font_family, size=11, bold=True, color="C00000")
                    
            # Status Fills
            if col_idx == 9: # Trạng thái duyệt
                val = str(cell.value)
                if val == "Đã duyệt":
                    cell.fill = fill_approved
                    cell.font = Font(name=font_family, size=11, bold=True, color="375623")
                elif val == "Từ chối":
                    cell.fill = fill_rejected
                    cell.font = Font(name=font_family, size=11, bold=True, color="C00000")
                else: # Chưa duyệt
                    cell.fill = fill_pending
                    cell.font = Font(name=font_family, size=11, bold=False, color="7F6000")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        
        # Don't auto-widen requirements & reason columns to infinity
        if col[0].column in [5, 8]: # Nhu Cầu & Lý Do
            ws.column_dimensions[col_letter].width = 45
            continue
            
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Write to memory file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Bao_Cao_Khach_Hang_Tiem_Nang.xlsx"
    )

if __name__ == '__main__':
    # Initial setup
    load_leads()
    print("Flask Server running on http://127.0.0.1:5000")
    app.run(host='127.0.0.1', port=5000, debug=True)

import os
import re
import io
import json
import sys
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Configure terminal output encoding to support Vietnamese characters on Windows
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass


# Configuration
SHEET_URL = "https://docs.google.com/spreadsheets/d/1hRvHE6RXm1peVG07avfApPEHocOcPld9IA94hE3vUGE/export?format=csv"
OUTPUT_FILE = "Bao_Cao_Khach_Hang_Tiem_Nang.xlsx"
MOCK_FILE = "mock_leads.json"

def rule_based_scorer(requirement):
    """
    Chấm điểm khách hàng tiềm năng dựa trên bộ quy tắc nghiệp vụ trong tieu_chi_cham_diem.txt.
    Không sử dụng bất kỳ API Key nào, phân tích dựa trên từ khóa và ngữ cảnh tối ưu.
    """
    if not requirement or not isinstance(requirement, str):
        return {
            "score": 0,
            "classification": "Trung bình",
            "reason": "Nhu cầu trống hoặc không hợp lệ"
        }
        
    req_lower = requirement.lower()
    score = 0
    reason_parts = []
    
    # 1. Trích xuất ngân sách (tỷ)
    budget_val = None
    # Khớp các mẫu: 35 tỷ, 35.5 tỷ, 1,5 tỷ, 1-2 tỷ, 3-4 tỷ
    matches = re.findall(r'(\d+(?:[.,]\d+)?)\s*(?:-|đến)?\s*(\d+(?:[.,]\d+)?)\s*(?:tỷ|ty)', req_lower)
    if matches:
        for match in matches:
            val1 = float(match[0].replace(',', '.'))
            val2 = float(match[1].replace(',', '.'))
            budget_val = max(val1, val2)
    else:
        single_matches = re.findall(r'(\d+(?:[.,]\d+)?)\s*(?:tỷ|ty)', req_lower)
        if single_matches:
            budget_val = max([float(m.replace(',', '.')) for m in single_matches])
            
    # 2. KIỂM TRA TIÊU CHÍ VIP (+50)
    vip_detected = False
    vip_reasons = []
    
    if budget_val is not None and budget_val >= 20.0:
        vip_detected = True
        vip_reasons.append(f"Ngân sách lớn ({budget_val} tỷ)")
    elif "tài chính mạnh" in req_lower or "tai chinh manh" in req_lower or "không thành vấn đề" in req_lower or "khong thanh van de" in req_lower:
        vip_detected = True
        vip_reasons.append("Tài chính mạnh")
        
    premium_types = ["biệt thự", "biet thu", "penthouse", "shophouse mặt đường", "shophouse mat duong", "đất công nghiệp", "dat cong nghiep", "sàn văn phòng diện tích lớn", "san van phong dien tich lon", "sàn văn phòng", "san van phong"]
    for pt in premium_types:
        if pt in req_lower:
            vip_detected = True
            vip_reasons.append(f"Loại hình cao cấp ({pt})")
            break
            
    prime_locations = ["quận 1", "quan 1", "ven sông", "ven song", "ocean park", "phú mỹ hưng", "phu my hung"]
    has_low_budget = (budget_val is not None and budget_val <= 2.0)
    if not has_low_budget:
        for pl in prime_locations:
            if pl in req_lower:
                vip_detected = True
                vip_reasons.append(f"Vị trí đắc địa ({pl})")
                break
            
    profiles = ["chủ doanh nghiệp", "chu doanh nghiep", "nhà đầu tư chuyên nghiệp", "nha dau tu chuyen nghiep", "mua sỉ", "mua si", "số lượng lớn", "so luong lon"]
    for pr in profiles:
        if pr in req_lower:
            vip_detected = True
            vip_reasons.append(f"Đối tượng VIP ({pr})")
            break
            
    urgency = ["pháp lý chuẩn", "phap ly chuan", "sổ hồng riêng", "so hong rieng", "gặp trực tiếp chủ đầu tư", "gap truc tiep chu dau tu"]
    for urg in urgency:
        if urg in req_lower:
            vip_detected = True
            vip_reasons.append(f"Tính cấp thiết & Minh bạch ({urg})")
            break
            
    # 3. KIỂM TRA TIÊU CHÍ RÁC (-50)
    trash_detected = False
    trash_reasons = []
    
    # Yêu cầu phi thực tế (nhà Q1 / trung tâm giá <= 2 tỷ)
    has_central = any(x in req_lower for x in ["quận 1", "quan 1", "trung tâm", "trung tam"])
    if has_central and budget_val is not None and budget_val <= 2.0:
        trash_detected = True
        trash_reasons.append(f"Yêu cầu phi thực tế (nhà trung tâm giá {budget_val} tỷ)")
    elif "vài trăm triệu" in req_lower or "vai tram trieu" in req_lower:
        trash_detected = True
        trash_reasons.append("Yêu cầu phi thực tế (biệt thự/nhà trung tâm vài trăm triệu)")
        
    no_demands = ["nhầm số", "nham so", "không có nhu cầu", "khong co nhu cau", "dữ liệu cũ", "du lieu cu", "nhầm ngành", "nham nganh", "nhầm người", "nham nguoi"]
    for nd in no_demands:
        if nd in req_lower:
            trash_detected = True
            trash_reasons.append(f"Không có nhu cầu ({nd})")
            break
            
    uncooperative = ["hỏi giá cho vui", "hoi gia cho vui", "chưa có ý định mua", "chua co y dinh mua", "không thiện chí", "thái độ không hợp tác", "khong hop tac"]
    for uc in uncooperative:
        if uc in req_lower:
            trash_detected = True
            trash_reasons.append(f"Khách hàng không thiện chí ({uc})")
            break
            
    # Kiểm tra Spam/Quảng cáo (chỉ kiểm tra khi không có ý định mua thực tế)
    has_buyer_intent = any(x in req_lower for x in ["cần mua", "can mua", "tìm mua", "tim mua", "mua chung cư", "mua chung cu", "mua nhà", "mua nha", "tìm căn", "tim can", "nhu cầu tìm"])
    if not has_buyer_intent:
        spam_services = ["bảo hiểm", "bao hiem", "dịch vụ vay", "dich vu vay", "cho vay", "mời chào dịch vụ", "moi chao dich vu", "chuyên cung cấp", "chuyen cung cap"]
        for ss in spam_services:
            if ss in req_lower:
                trash_detected = True
                trash_reasons.append(f"Spam/Quảng cáo dịch vụ ({ss})")
                break
                
    comm_errors = ["thuê bao", "thue bao", "không bắt máy", "khong bat may", "không phản hồi zalo", "khong phan hoi zalo", "không liên lạc được", "khong lien lac duoc"]
    for ce in comm_errors:
        if ce in req_lower:
            trash_detected = True
            trash_reasons.append(f"Lỗi liên lạc ({ce})")
            break
            
    # 4. TỔNG HỢP KẾT QUẢ CHẤM ĐIỂM
    if trash_detected:
        score = -50
        classification = "Rác"
        reason = "Trừ 50đ: " + ", ".join(trash_reasons)
    elif vip_detected:
        score = 50
        classification = "VIP"
        reason = "Cộng 50đ: " + ", ".join(vip_reasons)
    else:
        score = 0
        classification = "Trung bình"
        reason = "Khách hàng tầm trung, có nhu cầu thực tế liên quan đến chung cư/nhà phố hoặc cần tư vấn thêm."
        
    return {
        "score": score,
        "classification": classification,
        "reason": f"[Tự động] {reason}"
    }

def normalize_dataframe(df):
    """Chuẩn hóa tiêu đề cột từ Google Sheets thành cấu trúc dữ liệu chung."""
    cols = [str(c).lower().strip() for c in df.columns]
    
    name_cols = ['họ tên', 'ho ten', 'họ và tên', 'ho va ten', 'tên', 'ten', 'name', 'khách hàng', 'khach hang']
    phone_cols = ['số điện thoại', 'so dien thoai', 'sđt', 'sdt', 'phone', 'điện thoại', 'dien thoai', 'tel']
    email_cols = ['email', 'mail', 'thư điện tử']
    req_cols = ['nhu cầu chi tiết', 'nhu cau chi tiet', 'nhu cầu', 'nhu cau', 'chi tiết', 'chi tiet', 'nội dung', 'noi dung', 'notes', 'mô tả', 'mo ta']
    
    mapped_name = None
    mapped_phone = None
    mapped_email = None
    mapped_req = None
    
    for idx, c in enumerate(cols):
        orig_col = df.columns[idx]
        if c in name_cols and not mapped_name:
            mapped_name = orig_col
        elif c in phone_cols and not mapped_phone:
            mapped_phone = orig_col
        elif c in email_cols and not mapped_email:
            mapped_email = orig_col
        elif c in req_cols and not mapped_req:
            mapped_req = orig_col
            
    if not mapped_req:
        for col in df.columns:
            if df[col].dtype == object and df[col].str.len().mean() > 35:
                mapped_req = col
                break
        if not mapped_req and len(df.columns) >= 3:
            mapped_req = df.columns[2]
            
    if not mapped_name and len(df.columns) >= 1:
        mapped_name = df.columns[0]
        
    if not mapped_phone and len(df.columns) >= 2:
        mapped_phone = df.columns[1]

    leads = []
    for idx, row in df.iterrows():
        name = str(row[mapped_name]).strip() if mapped_name and pd.notna(row[mapped_name]) else "Khách hàng"
        phone = str(row[mapped_phone]).strip() if mapped_phone and pd.notna(row[mapped_phone]) else ""
        email = str(row[mapped_email]).strip() if mapped_email and pd.notna(row[mapped_email]) else ""
        requirement = str(row[mapped_req]).strip() if mapped_req and pd.notna(row[mapped_req]) else ""
        
        if not phone and not requirement:
            continue
            
        leads.append({
            "id": idx + 1,
            "name": name,
            "phone": phone,
            "email": email,
            "requirement": requirement
        })
    return leads

def fetch_data():
    """Tải dữ liệu từ Google Sheets hoặc sử dụng dữ liệu giả lập cục bộ làm dự phòng."""
    print(f"Đang tải dữ liệu từ Google Sheets: {SHEET_URL}...")
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    try:
        res = requests.get(SHEET_URL, headers=headers, timeout=15)
        if res.status_code == 200:
            df = pd.read_csv(io.StringIO(res.text))
            print(f"SUCCESS: Tải thành công {len(df)} dòng dữ liệu từ Google Sheets.")
            return normalize_dataframe(df)
        else:
            print(f"WARNING: Google Sheets trả về mã lỗi {res.status_code} (Bảng tính có thể ở chế độ riêng tư).")
    except Exception as e:
        print(f"WARNING: Không thể kết nối với Google Sheets. Lỗi: {str(e)}")
        
    # Thử nạp dữ liệu mẫu dự phòng cục bộ
    print("-> Đang thử nạp dữ liệu từ tệp mock_leads.json dự phòng...")
    if os.path.exists(MOCK_FILE):
        try:
            with open(MOCK_FILE, 'r', encoding='utf-8') as f:
                leads_data = json.load(f)
                print(f"SUCCESS: Đã nạp {len(leads_data)} khách hàng từ {MOCK_FILE}.")
                return leads_data
        except Exception as e:
            print(f"ERROR: Lỗi đọc tệp {MOCK_FILE}: {str(e)}")
            
    # Tạo dữ liệu giả lập mẫu mặc định
    print("-> Đang tạo dữ liệu mẫu mặc định...")
    default_data = [
        {"id": 1, "name": "Nguyễn Văn Hưng", "phone": "0912345678", "email": "hung@email.com", "requirement": "Tôi là chủ doanh nghiệp cần tìm mua biệt thự đơn lập Vinhomes Ocean Park, tài chính tầm 35 tỷ."},
        {"id": 2, "name": "Trần Thị Lan", "phone": "0987654321", "email": "lan@email.com", "requirement": "Mua sỉ diện tích lớn sàn văn phòng Quận 1 ven sông. Tài chính mạnh không thành vấn đề."},
        {"id": 3, "name": "Lê Hoàng Nam", "phone": "0915112233", "email": "nam@email.com", "requirement": "Mua chung cư 2 phòng ngủ giá 3-4 tỷ Quận 7 cho gia đình. Cần vay ngân hàng hỗ trợ."},
        {"id": 4, "name": "Hoàng Văn Tèo", "phone": "0966554433", "email": "teo@email.com", "requirement": "Cần mua nhà mặt phố trung tâm Quận 1 có hồ bơi giá 1.5 tỷ, yêu cầu sổ hồng riêng."},
        {"id": 5, "name": "Bảo Hiểm ABC", "phone": "0909998877", "email": "spam@email.com", "requirement": "Chào bạn, bên mình chuyên cung cấp bảo hiểm sức khỏe cá nhân..."}
    ]
    return default_data

def export_to_excel(leads):
    """Xuất danh sách khách hàng đã chấm điểm ra file Excel được định dạng màu sắc đẹp mắt."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo Chấm Điểm Leads"
    
    # Bật hiển thị grid lines
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "Mã Lead", "Họ Tên Khách Hàng", "Số Điện Thoại", "Email", 
        "Nhu Cầu Chi Tiết", "Điểm Đánh Giá", "Phân Loại", "Chi Tiết Đánh Giá"
    ]
    ws.append(headers)
    
    # Định dạng
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Xanh dương đậm
    data_font = Font(name=font_family, size=11)
    
    thin = Side(border_style="thin", color="D9D9D9")
    double = Side(border_style="double", color="1F4E79")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Màu sắc phân loại
    fill_vip = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Vàng kim nhạt
    fill_med = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Xanh nhạt
    fill_trash = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Đỏ nhạt
    
    # Định dạng tiêu đề cột
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = Border(left=thin, right=thin, top=thin, bottom=double)
    ws.row_dimensions[1].height = 28
    
    # Điền dữ liệu
    for lead in leads:
        row_data = [
            f"L{lead.get('id', 0):03d}",
            lead.get('name', ''),
            lead.get('phone', ''),
            lead.get('email', ''),
            lead.get('requirement', ''),
            lead.get('score', 0),
            lead.get('classification', 'Trung bình'),
            lead.get('reason', '')
        ]
        ws.append(row_data)
        
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = 24
        
        # Định dạng từng ô dữ liệu
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = grid_border
            cell.alignment = align_left
            
            # Căn giữa một số cột
            if col_idx in [1, 3, 6, 7]:
                cell.alignment = align_center
                
            # Tô màu phân loại
            if col_idx == 7:
                val = str(cell.value)
                if val == "VIP":
                    cell.fill = fill_vip
                    cell.font = Font(name=font_family, size=11, bold=True, color="7F6000")
                elif val == "Trung bình":
                    cell.fill = fill_med
                    cell.font = Font(name=font_family, size=11, bold=False, color="1F4E79")
                elif val == "Rác":
                    cell.fill = fill_trash
                    cell.font = Font(name=font_family, size=11, bold=True, color="C00000")
                    
    # Tự động căn chỉnh độ rộng cột
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        
        # Cột nhu cầu & lý do cho cố định độ rộng để dễ đọc
        if col[0].column in [5, 8]:
            ws.column_dimensions[col_letter].width = 45
            continue
            
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(OUTPUT_FILE)
    print(f"SUCCESS: Báo cáo đã được xuất thành công ra tệp '{OUTPUT_FILE}'.")

def main():
    print("=========================================")
    print("BẮT ĐẦU QUY TRÌNH CHẤM ĐIỂM LEADS TỰ ĐỘNG")
    print("=========================================")
    
    # 1. Fetch
    leads_list = fetch_data()
    
    # 2. Score
    print("\nĐang chấm điểm tự động theo quy tắc...")
    for lead in leads_list:
        req = lead.get("requirement", "")
        res = rule_based_scorer(req)
        
        lead["score"] = res["score"]
        lead["classification"] = res["classification"]
        lead["reason"] = res["reason"]
        
        # Print results in terminal (ASCII only to avoid encoding errors)
        clean_name = lead.get('name', '').encode('ascii', errors='replace').decode('ascii')
        print(f"- Lead {lead.get('id', 0):02d} | Khách hàng: {clean_name:<15} | Điểm: {res['score']:4d} | Phân loại: {res['classification']}")

    # 3. Export
    print("\nĐang định dạng và ghi tệp Excel...")
    export_to_excel(leads_list)
    print("=========================================")
    print("QUY TRÌNH HOÀN THÀNH XUẤT SẮC!")
    print("=========================================")

if __name__ == "__main__":
    main()

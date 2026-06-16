import streamlit as st
import pandas as pd
import openpyxl
import io
import requests
import json
import os
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Page Configuration
st.set_page_config(
    page_title="AI Lead Scoring Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Constants
SHEET_URL = "https://docs.google.com/spreadsheets/d/1hRvHE6RXm1peVG07avfApPEHocOcPld9IA94hE3vUGE/export?format=csv"
OUTPUT_FILE = "Bao_Cao_Khach_Hang_Tiem_Nang.xlsx"

# ----------------------------------------------------
# 1. CORE BUSINESS RULES (RULE-BASED LEAD SCORER)
# ----------------------------------------------------
def rule_based_scorer(requirement):
    """
    Chấm điểm khách hàng tiềm năng dựa trên bộ quy tắc nghiệp vụ trong tieu_chi_cham_diem.txt.
    Không sử dụng bất kỳ API Key nào.
    
    Quy tắc điểm:
    - Điểm gốc mặc định: 50đ (Tiềm năng trung bình)
    - Thỏa VIP: Cộng 50đ -> Đạt 100đ (Khách VIP)
    - Thỏa Rác: Trừ 50đ -> Về 0đ (Không tiềm năng)
    """
    if not requirement or not isinstance(requirement, str):
        return {
            "score": 50,
            "classification": "Ấm",
            "reason": "Nhu cầu trống hoặc không hợp lệ"
        }
        
    req_lower = requirement.lower()
    score = 50 # Điểm gốc
    reason_parts = []
    
    # 1. Trích xuất ngân sách (tỷ)
    budget_val = None
    matches = re.findall(r'(\d+(?:[.,]\d+)?)\s*(?:-|đến)?\s*(\d+(?:[.,]\d+)?)\s*(?:tỷ|ty)', req_lower)
    if matches:
        for match in matches:
            val1 = float(match[0].replace(',', '.'))
            val2 = float(match[1].replace(',', '.'))
            budget_val = max(val1, val2)
    else:
        single_matches = re.findall(r'(\d+(?:[.,]\d+)?)\s*(?:tỷ|ty)', req_lower)
        if single_matches:
            budget_val = max([float(m.replace(',', '.')) for m in single_matches])
            
    # 2. KIỂM TRA TIÊU CHÍ VIP (+50 -> Đạt 100đ)
    vip_detected = False
    vip_reasons = []
    
    if budget_val is not None and budget_val >= 20.0:
        vip_detected = True
        vip_reasons.append(f"Ngân sách lớn ({budget_val} tỷ)")
    elif "tài chính mạnh" in req_lower or "tai chinh manh" in req_lower or "không thành vấn đề" in req_lower or "khong thanh van de" in req_lower:
        vip_detected = True
        vip_reasons.append("Tài chính mạnh")
        
    premium_types = ["biệt thự", "biet thu", "penthouse", "shophouse mặt đường", "shophouse mat duong", "đất công nghiệp", "dat cong nghiep", "sàn văn phòng diện tích lớn", "san van phong dien tich lon", "sàn văn phòng", "san van phong"]
    for pt in premium_types:
        if pt in req_lower:
            vip_detected = True
            vip_reasons.append(f"Loại hình cao cấp ({pt})")
            break
            
    prime_locations = ["quận 1", "quan 1", "ven sông", "ven song", "ocean park", "phú mỹ hưng", "phu my hung"]
    has_low_budget = (budget_val is not None and budget_val <= 2.0)
    if not has_low_budget:
        for pl in prime_locations:
            if pl in req_lower:
                vip_detected = True
                vip_reasons.append(f"Vị trí đắc địa ({pl})")
                break
            
    profiles = ["chủ doanh nghiệp", "chu doanh nghiep", "nhà đầu tư chuyên nghiệp", "nha dau tu chuyen nghiep", "mua sỉ", "mua si", "số lượng lớn", "so luong lon"]
    for pr in profiles:
        if pr in req_lower:
            vip_detected = True
            vip_reasons.append(f"Đối tượng VIP ({pr})")
            break
            
    urgency = ["pháp lý chuẩn", "phap ly chuan", "sổ hồng riêng", "so hong rieng", "gặp trực tiếp chủ đầu tư", "gap truc tiep chu dau tu"]
    for urg in urgency:
        if urg in req_lower:
            vip_detected = True
            vip_reasons.append(f"Tính cấp thiết & Minh bạch ({urg})")
            break
            
    # 3. KIỂM TRA TIÊU CHÍ RÁC (-50 -> Về 0đ)
    trash_detected = False
    trash_reasons = []
    
    # Yêu cầu phi thực tế (nhà Q1 / trung tâm giá <= 2 tỷ)
    has_central = any(x in req_lower for x in ["quận 1", "quan 1", "trung tâm", "trung tam"])
    if has_central and budget_val is not None and budget_val <= 2.0:
        trash_detected = True
        trash_reasons.append(f"Yêu cầu phi thực tế (nhà trung tâm giá {budget_val} tỷ)")
    elif "vài trăm triệu" in req_lower or "vai tram trieu" in req_lower:
        trash_detected = True
        trash_reasons.append("Yêu cầu phi thực tế (biệt thự/nhà trung tâm vài trăm triệu)")
        
    no_demands = ["nhầm số", "nham so", "không có nhu cầu", "khong co nhu cau", "dữ liệu cũ", "du lieu cu", "nhầm ngành", "nham nganh", "nhầm người", "nham nguoi"]
    for nd in no_demands:
        if nd in req_lower:
            trash_detected = True
            trash_reasons.append(f"Không có nhu cầu ({nd})")
            break
            
    uncooperative = ["hỏi giá cho vui", "hoi gia cho vui", "chưa có ý định mua", "chua co y dinh mua", "không thiện chí", "thái độ không hợp tác", "khong hop tac"]
    for uc in uncooperative:
        if uc in req_lower:
            trash_detected = True
            trash_reasons.append(f"Khách hàng không thiện chí ({uc})")
            break
            
    # Kiểm tra Spam/Quảng cáo
    has_buyer_intent = any(x in req_lower for x in ["cần mua", "can mua", "tìm mua", "tim mua", "mua chung cư", "mua chung cu", "mua nhà", "mua nha", "tìm căn", "tim can", "nhu cầu tìm"])
    if not has_buyer_intent:
        spam_services = ["bảo hiểm", "bao hiem", "dịch vụ vay", "dich vu vay", "cho vay", "mời chào dịch vụ", "moi chao dich vu", "chuyên cung cấp", "chuyen cung cap"]
        for ss in spam_services:
            if ss in req_lower:
                trash_detected = True
                trash_reasons.append(f"Spam/Quảng cáo dịch vụ ({ss})")
                break
                
    comm_errors = ["thuê bao", "thue bao", "không bắt máy", "khong bat may", "không phản hồi zalo", "khong phan hoi zalo", "không liên lạc được", "khong lien lac duoc"]
    for ce in comm_errors:
        if ce in req_lower:
            trash_detected = True
            trash_reasons.append(f"Lỗi liên lạc ({ce})")
            break
            
    # 4. TỔNG HỢP ĐIỂM SỐ & PHÂN LOẠI
    if trash_detected:
        score = 0
        classification = "Rác"
        reason = "Trừ 50đ: " + ", ".join(trash_reasons)
    elif vip_detected:
        score = 100
        classification = "Nóng"
        reason = "Cộng 50đ: " + ", ".join(vip_reasons)
    else:
        score = 50
        classification = "Ấm"
        reason = "Khách hàng tầm trung, có nhu cầu thực tế liên quan đến chung cư/nhà phố hoặc cần tư vấn thêm."
        
    return {
        "score": score,
        "classification": classification,
        "reason": f"[Tự động] {reason}"
    }

def normalize_dataframe(df):
    """Chuẩn hóa tiêu đề cột từ Google Sheets thành cấu trúc dữ liệu chung."""
    cols = [str(c).lower().strip() for c in df.columns]
    
    name_cols = ['họ tên', 'ho ten', 'họ và tên', 'ho va ten', 'tên', 'ten', 'name', 'khách hàng', 'khach hang']
    phone_cols = ['số điện thoại', 'so dien thoai', 'sđt', 'sdt', 'phone', 'điện thoại', 'dien thoai', 'tel']
    email_cols = ['email', 'mail', 'thư điện tử']
    req_cols = ['nhu cầu chi tiết', 'nhu cau chi tiet', 'nhu cầu', 'nhu cau', 'chi tiết', 'chi tiet', 'nội dung', 'noi dung', 'notes', 'mô tả', 'mo ta']
    
    mapped_name = None
    mapped_phone = None
    mapped_email = None
    mapped_req = None
    
    for idx, c in enumerate(cols):
        orig_col = df.columns[idx]
        if c in name_cols and not mapped_name:
            mapped_name = orig_col
        elif c in phone_cols and not mapped_phone:
            mapped_phone = orig_col
        elif c in email_cols and not mapped_email:
            mapped_email = orig_col
        elif c in req_cols and not mapped_req:
            mapped_req = orig_col
            
    if not mapped_req:
        for col in df.columns:
            if df[col].dtype == object and df[col].str.len().mean() > 35:
                mapped_req = col
                break
        if not mapped_req and len(df.columns) >= 3:
            mapped_req = df.columns[2]
            
    if not mapped_name and len(df.columns) >= 1:
        mapped_name = df.columns[0]
        
    if not mapped_phone and len(df.columns) >= 2:
        mapped_phone = df.columns[1]

    leads = []
    for idx, row in df.iterrows():
        name = str(row[mapped_name]).strip() if mapped_name and pd.notna(row[mapped_name]) else "Khách hàng"
        phone = str(row[mapped_phone]).strip() if mapped_phone and pd.notna(row[mapped_phone]) else ""
        email = str(row[mapped_email]).strip() if mapped_email and pd.notna(row[mapped_email]) else ""
        requirement = str(row[mapped_req]).strip() if mapped_req and pd.notna(row[mapped_req]) else ""
        
        if not phone and not requirement:
            continue
            
        leads.append({
            "id": idx + 1,
            "name": name,
            "phone": phone,
            "email": email,
            "requirement": requirement,
            "score": None,
            "classification": None,
            "reason": None,
            "status": "Chờ duyệt",
            "reviewer_notes": ""
        })
    return leads

# ----------------------------------------------------
# 2. SESSION STATE & DEFAULT MOCK DATA
# ----------------------------------------------------
if 'leads' not in st.session_state:
    st.session_state.leads = []
if 'processed_count' not in st.session_state:
    st.session_state.processed_count = 0

# Mock leads for fallback
MOCK_DATA = [
  {
    "id": 1,
    "name": "Nguyễn Văn Hưng",
    "phone": "0912345678",
    "email": "hung.nguyen@email.com",
    "requirement": "Tôi là chủ doanh nghiệp, đang có nhu cầu tìm mua một căn biệt thự đơn lập tại Vinhomes Ocean Park. Ngân sách khoảng 35 tỷ đồng, tài chính đã sẵn sàng thanh toán ngay. Yêu cầu pháp lý chuẩn 100%, có sổ hồng riêng và muốn gặp trực tiếp chủ đầu tư để thương thảo.",
    "status": "Chờ duyệt",
    "reviewer_notes": ""
  },
  {
    "id": 2,
    "name": "Trần Thị Lan",
    "phone": "0987654321",
    "email": "lan.tran@email.com",
    "requirement": "Đại diện quỹ đầu tư mua sỉ diện tích lớn sàn văn phòng tại Quận 1 ven sông Sài Gòn để làm trụ sở. Tài chính mạnh không thành vấn đề, yêu cầu pháp lý hoàn chỉnh sổ hồng trao tay.",
    "status": "Chờ duyệt",
    "reviewer_notes": ""
  },
  {
    "id": 3,
    "name": "Phạm Minh Đức",
    "phone": "0903334455",
    "email": "duc.pham@email.com",
    "requirement": "Cần tìm căn Penthouse cao cấp tại Phú Mỹ Hưng. Yêu cầu có tầm nhìn ven sông thông thoáng, tài chính tầm 25 tỷ thanh toán nhanh gọn. Muốn làm việc trực tiếp với chủ đầu tư để thương lượng giá.",
    "status": "Chờ duyệt",
    "reviewer_notes": ""
  },
  {
    "id": 4,
    "name": "Lê Hoàng Nam",
    "phone": "0915112233",
    "email": "nam.le@email.com",
    "requirement": "Cần mua chung cư 2 phòng ngủ giá từ 3-4 tỷ đồng tại khu vực Quận 7 hoặc Nhà Bè cho gia đình ở. Cần dự án có chính sách hỗ trợ vay vốn ngân hàng lãi suất tốt.",
    "status": "Chờ duyệt",
    "reviewer_notes": ""
  },
  {
    "id": 5,
    "name": "Vũ Thị Hồng",
    "phone": "0978223344",
    "email": "hong.vu@email.com",
    "requirement": "Tìm nhà phố khu vực Quận 9 hoặc Thủ Đức tầm giá 6-7 tỷ. Cần tư vấn thêm về pháp lý và quy hoạch lộ giới đường trước nhà.",
    "status": "Chờ duyệt",
    "reviewer_notes": ""
  },
  {
    "id": 6,
    "name": "Hoàng Văn Tèo",
    "phone": "0966554433",
    "email": "teo.hoang@email.com",
    "requirement": "Cần mua nhà mặt phố trung tâm Quận 1 có sân vườn và hồ bơi, giá tầm 1.5 tỷ trở lại. Đầy đủ sổ hồng riêng.",
    "status": "Chờ duyệt",
    "reviewer_notes": ""
  },
  {
    "id": 7,
    "name": "Đỗ Minh Quân",
    "phone": "0933889900",
    "email": "quan.do@email.com",
    "requirement": "Nhầm số rồi em ơi, anh không có nhu cầu mua bất động sản đâu nhé.",
    "status": "Chờ duyệt",
    "reviewer_notes": ""
  },
  {
    "id": 8,
    "name": "Bảo Hiểm ABC",
    "phone": "0909998877",
    "email": "info@insurance.com",
    "requirement": "Chào bạn, bên mình chuyên cung cấp các gói bảo hiểm sức khỏe ưu việt cho doanh nghiệp và cá nhân, chiết khấu lên đến 30%...",
    "status": "Chờ duyệt",
    "reviewer_notes": ""
  },
  {
    "id": 9,
    "name": "Nguyễn Thu Trang",
    "phone": "0911002233",
    "email": "trang.nguyen@email.com",
    "requirement": "Số thuê bao không liên lạc được, đã gọi nhiều lần vào các khung giờ khác nhau nhưng không bắt máy, nhắn tin Zalo không thấy phản hồi.",
    "status": "Chờ duyệt",
    "reviewer_notes": ""
  }
]

# Set initial mock if empty
if not st.session_state.leads:
    st.session_state.leads = MOCK_DATA
    # Run initial scoring
    for l in st.session_state.leads:
        res = rule_based_scorer(l.get('requirement', ''))
        l['score'] = res['score']
        l['classification'] = res['classification']
        l['reason'] = res['reason']
    st.session_state.processed_count = len(MOCK_DATA)

# ----------------------------------------------------
# 3. INTERACTIVE STYLING (Dark theme CSS)
# ----------------------------------------------------
st.markdown("""
<style>
    /* Metric Card Styling */
    .custom-card-container {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 1rem;
        margin-bottom: 2rem;
        margin-top: 1rem;
    }
    .custom-card {
        background-color: #0E1325;
        border: 1px solid #1E293B;
        border-radius: 12px;
        padding: 1.25rem;
        text-align: center;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
    }
    .custom-card .value {
        font-size: 2.2rem;
        font-weight: 800;
        font-family: 'Outfit', sans-serif;
        margin-bottom: 0.25rem;
    }
    .custom-card.total .value { color: #3B82F6; }
    .custom-card.vip .value { color: #10B981; }
    .custom-card.medium .value { color: #F59E0B; }
    .custom-card.trash .value { color: #EF4444; }
    
    .custom-card .label {
        font-size: 0.8rem;
        font-weight: 600;
        color: #94A3B8;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# 4. SIDEBAR - System configuration & filters
# ----------------------------------------------------
with st.sidebar:
    st.markdown("### ⚙️ Cấu hình hệ thống")
    data_source = st.selectbox(
        "Nguồn dữ liệu đầu vào",
        ["Google Sheets Link", "Tải tệp từ máy (.xlsx, .csv)"],
        help="Chọn phương thức nhập dữ liệu khách hàng cần chấm điểm"
    )
    
    if data_source == "Google Sheets Link":
        sheet_url_input = st.text_input(
            "Đường dẫn Google Sheets (CSV Export)", 
            value=SHEET_URL,
            help="Đường dẫn trích xuất định dạng CSV của Google Sheets"
        )
    else:
        uploaded_file = st.file_uploader(
            "Chọn tệp khách hàng từ máy",
            type=["csv", "xlsx", "xls"],
            help="Hỗ trợ định dạng .csv, .xlsx, .xls"
        )
    
    st.markdown("### 🔍 Bộ lọc hiển thị")
    search_q = st.text_input("Tìm kiếm theo Tên / Số điện thoại", value="")
    
    st.markdown("---")
    st.markdown("### 💡 Quy tắc chấm điểm chính:")
    st.markdown("""
- **Cộng 50đ (Khách VIP - Đạt 100đ):** Ngân sách $\ge$ 20 tỷ; Tìm biệt thự đơn lập, penthouse, shophouse mặt đường lớn, quỹ đất lớn; Vị trí đắc địa (Q1, ven sông, Phú Mỹ Hưng...); Yêu cầu pháp lý 100%, gặp trực tiếp CĐT.
- **Trừ 50đ (Khách Rác - Về 0đ):** Yêu cầu phi thực tế (giá rẻ vô lý...); Không có nhu cầu...; Spam/Quảng cáo; Thông tin liên lạc lỗi...
""")

# ----------------------------------------------------
# 5. MAIN PANEL - Title, sync, stats
# ----------------------------------------------------
st.title("AI LEAD SCORING DASHBOARD")
st.caption("Bảng điều khiển phân tích & chấm điểm khách hàng tiềm năng tự động")

# Process Google Sheets Sync or File Upload
if data_source == "Google Sheets Link":
    if st.button("📥 Tải dữ liệu & Chấm điểm từ Google Sheet", type="primary"):
        headers = {'User-Agent': 'Mozilla/5.0'}
        csv_url = sheet_url_input
        if "/edit" in sheet_url_input:
            match = re.search(r'\/spreadsheets\/d\/([a-zA-Z0-9-_]+)', sheet_url_input)
            if match:
                sheet_id = match.group(1)
                gid = "0"
                gid_match = re.search(r'gid=(\d+)', sheet_url_input)
                if gid_match:
                    gid = gid_match.group(1)
                csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
                
        try:
            res = requests.get(csv_url, headers=headers, timeout=15)
            if res.status_code == 200:
                df = pd.read_csv(io.StringIO(res.text))
                normalized = normalize_dataframe(df)
                
                # Score
                for lead in normalized:
                    res_score = rule_based_scorer(lead.get('requirement', ''))
                    lead['score'] = res_score['score']
                    lead['classification'] = res_score['classification']
                    lead['reason'] = res_score['reason']
                    
                st.session_state.leads = normalized
                st.session_state.processed_count = len(normalized)
                st.success(f"Đã xử lý thành công {st.session_state.processed_count} dòng dữ liệu!")
            else:
                st.error(f"Google Sheet trả về mã lỗi {res.status_code} (Bảng tính riêng tư). Không thể đồng bộ.")
        except Exception as e:
            st.error(f"Lỗi đồng bộ: {str(e)}")
else:
    # File upload handling
    if 'last_uploaded_file' not in st.session_state:
        st.session_state.last_uploaded_file = None
        
    if uploaded_file is not None and uploaded_file != st.session_state.last_uploaded_file:
        try:
            if uploaded_file.name.endswith(".csv"):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
                
            normalized = normalize_dataframe(df)
            
            # Score
            for lead in normalized:
                res_score = rule_based_scorer(lead.get('requirement', ''))
                lead['score'] = res_score['score']
                lead['classification'] = res_score['classification']
                lead['reason'] = res_score['reason']
                
            st.session_state.leads = normalized
            st.session_state.processed_count = len(normalized)
            st.session_state.last_uploaded_file = uploaded_file
            st.success(f"Đã xử lý thành công {st.session_state.processed_count} dòng dữ liệu từ tệp cục bộ!")
        except Exception as e:
            st.error(f"Lỗi phân tích tệp: {str(e)}")

# Placeholders for layout sections (ensuring metrics & charts display above filters & table)
metrics_placeholder = st.container()
charts_placeholder = st.container()
filters_placeholder = st.container()
table_placeholder = st.container()

# Filter and Search processing
filtered_leads = st.session_state.leads

# Apply search
if search_q.strip() != "":
    q = search_q.lower().strip()
    filtered_leads = [
        l for l in filtered_leads 
        if q in l.get('name', '').lower() or q in l.get('phone', '').lower()
    ]

# Render filters in their designated placeholder
with filters_placeholder:
    st.markdown("---")
    st.markdown("### 📝 2. Bảng Kiểm Duyệt (Dành cho Kế Toán / Sales)")
    st.markdown("*Hệ thống đã tự động gán Từ khóa (Tags) và Gợi ý hành động. Bạn có thể dùng bộ lọc dưới đây để tìm và phê duyệt khách hàng nhanh chóng.*")
    
    st.markdown("#### 🔎 Bộ Lọc Dữ Liệu Thông Minh")
    filter_col1, filter_col2 = st.columns(2)
    with filter_col1:
        st.write("Lọc theo Phân loại AI:")
        selected_classes = st.multiselect(
            "Lọc theo Phân loại AI:",
            options=["Nóng", "Ấm", "Rác"],
            default=["Nóng", "Ấm", "Rác"],
            label_visibility="collapsed"
        )
    with filter_col2:
        st.write("Lọc theo Trạng thái duyệt:")
        selected_statuses = st.multiselect(
            "Lọc theo Trạng thái duyệt:",
            options=["Chờ duyệt", "Đã duyệt", "Loại bỏ"],
            default=["Chờ duyệt", "Đã duyệt", "Loại bỏ"],
            label_visibility="collapsed"
        )

# Apply main page smart filters
filtered_leads = [
    l for l in filtered_leads 
    if l.get('classification', 'Ấm') in selected_classes
]
filtered_leads = [
    l for l in filtered_leads 
    if l.get('status', 'Chờ duyệt') in selected_statuses
]

# Calculate metrics from all leads
all_leads_df = pd.DataFrame(st.session_state.leads)
total_leads = len(all_leads_df)
vip_leads = sum(1 for l in st.session_state.leads if l.get('classification') == 'Nóng')
medium_leads = sum(1 for l in st.session_state.leads if l.get('classification') == 'Ấm')
trash_leads = sum(1 for l in st.session_state.leads if l.get('classification') == 'Rác')

# Display Custom HTML Metric Cards inside metrics placeholder
with metrics_placeholder:
    metric_html = f"""
    <div class="custom-card-container">
        <div class="custom-card total">
            <div class="value">{total_leads}</div>
            <div class="label">Tổng khách hàng</div>
        </div>
        <div class="custom-card vip">
            <div class="value">{vip_leads}</div>
            <div class="label">Khách hàng Nóng</div>
        </div>
        <div class="custom-card medium">
            <div class="value">{medium_leads}</div>
            <div class="label">Khách hàng Ấm</div>
        </div>
        <div class="custom-card trash">
            <div class="value">{trash_leads}</div>
            <div class="label">Khách hàng Rác</div>
        </div>
    </div>
    """
    st.markdown(metric_html, unsafe_allow_html=True)

# Render charts inside charts placeholder
with charts_placeholder:
    st.markdown("### 📊 Biểu đồ phân tích trực quan")
    
    chart_col1, chart_col2 = st.columns(2)
    
    with chart_col1:
        st.markdown("##### 📈 Tỉ lệ phân loại Khách hàng")
        class_counts = {
            "Rác": sum(1 for l in filtered_leads if l.get('classification') == 'Rác'),
            "Ấm": sum(1 for l in filtered_leads if l.get('classification') == 'Ấm'),
            "Nóng": sum(1 for l in filtered_leads if l.get('classification') == 'Nóng')
        }
        df_chart_class = pd.DataFrame(
            list(class_counts.items()), 
            columns=['Phân loại', 'Số lượng']
        ).set_index('Phân loại')
        st.bar_chart(df_chart_class, use_container_width=True)

    with chart_col2:
        st.markdown("##### 📉 Phân bố điểm số tiềm năng")
        score_counts = {
            "0": sum(1 for l in filtered_leads if l.get('score') == 0),
            "50": sum(1 for l in filtered_leads if l.get('score') == 50),
            "100": sum(1 for l in filtered_leads if l.get('score') == 100)
        }
        df_chart_score = pd.DataFrame(
            list(score_counts.items()), 
            columns=['Điểm số', 'Số lượng']
        ).set_index('Điểm số')
        st.bar_chart(df_chart_score, use_container_width=True)

# Render table inside table placeholder
with table_placeholder:
    if filtered_leads:
        df_display = pd.DataFrame(filtered_leads)
        
        # Ensure missing columns are present
        for col in ['score', 'classification', 'reason', 'status', 'reviewer_notes']:
            if col not in df_display.columns:
                df_display[col] = None
                
        col_order = ['id', 'name', 'phone', 'requirement', 'score', 'classification', 'reason', 'status', 'reviewer_notes']
        df_display = df_display[col_order]
        
        # Interactive Data Editor
        edited_df = st.data_editor(
            df_display,
            column_config={
                "id": st.column_config.NumberColumn("Mã", disabled=True),
                "name": st.column_config.TextColumn("Họ tên", required=True),
                "phone": st.column_config.TextColumn("Số điện thoại"),
                "requirement": st.column_config.TextColumn("Nhu cầu", width="large"),
                "score": st.column_config.NumberColumn("Điểm AI", disabled=True),
                "classification": st.column_config.SelectboxColumn(
                    "Phân loại AI",
                    options=["Nóng", "Ấm", "Rác"],
                    disabled=True
                ),
                "reason": st.column_config.TextColumn("Lý do chấm điểm", disabled=True, width="medium"),
                "status": st.column_config.SelectboxColumn(
                    "Duyệt (Human)",
                    options=["Chờ duyệt", "Đã duyệt", "Loại bỏ"],
                    required=True
                ),
                "reviewer_notes": st.column_config.TextColumn("Ghi chú kiểm duyệt", width="medium")
            },
            disabled=["id", "score", "classification", "reason"],
            hide_index=True,
            use_container_width=True
        )
        
        # Sync edited data back to main session state
        for row in edited_df.to_dict('records'):
            # Update matching lead in main session state
            for lead in st.session_state.leads:
                if lead['id'] == row['id']:
                    lead['name'] = row['name']
                    lead['phone'] = row['phone']
                    lead['requirement'] = row['requirement']
                    lead['status'] = row['status']
                    lead['reviewer_notes'] = row['reviewer_notes']
                    break
                    
        st.markdown("---")
    
        # Excel Generation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo cáo Chấm Điểm Leads"
        ws.views.sheetView[0].showGridLines = True
        
        headers_excel = [
            "Mã Lead", "Họ Tên Khách Hàng", "Số Điện Thoại", "Email", 
            "Nhu Cầu Chi Tiết", "Điểm Đánh Giá", "Phân Loại", "Lý Do Đánh Giá", 
            "Trạng Thái Duyệt", "Ghi Chú Kiểm Duyệt"
        ]
        ws.append(headers_excel)
        
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        data_font = Font(name=font_family, size=11)
        
        thin = Side(border_style="thin", color="D9D9D9")
        double = Side(border_style="double", color="1F4E79")
        grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        fill_vip = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        fill_med = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        fill_trash = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        fill_approved = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_rejected = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        fill_pending = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        for col_idx, h in enumerate(headers_excel, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = Border(left=thin, right=thin, top=thin, bottom=double)
        ws.row_dimensions[1].height = 28
        
        for lead in st.session_state.leads:
            row_data = [
                f"L{lead.get('id', 0):03d}",
                lead.get('name', ''),
                lead.get('phone', ''),
                lead.get('email', ''),
                lead.get('requirement', ''),
                lead.get('score', 0) if lead.get('score') is not None else "",
                lead.get('classification', 'Ấm'),
                lead.get('reason', ''),
                lead.get('status', 'Chờ duyệt'),
                lead.get('reviewer_notes', '')
            ]
            ws.append(row_data)
            
            row_idx = ws.max_row
            ws.row_dimensions[row_idx].height = 24
            
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = grid_border
                cell.alignment = align_left
                
                if col_idx in [1, 3, 6, 7, 9]:
                    cell.alignment = align_center
                    
                if col_idx == 7:
                    val = str(cell.value)
                    if val == "Nóng":
                        cell.fill = fill_vip
                        cell.font = Font(name=font_family, size=11, bold=True, color="7F6000")
                    elif val == "Ấm":
                        cell.fill = fill_med
                        cell.font = Font(name=font_family, size=11, bold=False, color="1F4E79")
                    elif val == "Rác":
                        cell.fill = fill_trash
                        cell.font = Font(name=font_family, size=11, bold=True, color="C00000")
                        
                if col_idx == 9:
                    val = str(cell.value)
                    if val == "Đã duyệt":
                        cell.fill = fill_approved
                        cell.font = Font(name=font_family, size=11, bold=True, color="375623")
                    elif val == "Loại bỏ":
                        cell.fill = fill_rejected
                        cell.font = Font(name=font_family, size=11, bold=True, color="C00000")
                    else:
                        cell.fill = fill_pending
                        cell.font = Font(name=font_family, size=11, bold=False, color="7F6000")

        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            if col[0].column in [5, 8]:
                ws.column_dimensions[col_letter].width = 40
                continue
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        output_excel = io.BytesIO()
        wb.save(output_excel)
        output_excel.seek(0)
        
        st.download_button(
            label="📥 Tải xuống File Excel (.xlsx)",
            data=output_excel.getvalue(),
            file_name="Bao_Cao_Khach_Hang_Tiem_Nang.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    else:
        st.info("Không tìm thấy kết quả phù hợp với bộ lọc hiển thị.")
